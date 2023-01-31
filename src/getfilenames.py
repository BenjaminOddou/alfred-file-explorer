#!/usr/bin/env python3

import os
import re
import datetime
import openpyxl
import subprocess

# Get the current date and time as a datetime object
now = datetime.datetime.now()

# Use the strftime() method to format the datetime object as a string
formatted_time = now.strftime("%d-%m-%Y_%H:%M:%S")

# Environment variables
links_list = os.environ['_links_list'] # List of folder/files
temp_file = os.environ['data_folder'] + '/tmp_' + formatted_time + '.xlsx'
depth = int(os.environ['depth'])
reject_regex = os.environ['reject_regex']
reject_bool = bool(reject_regex)
reject_pattern = re.compile(r'{}'.format(reject_regex))

# Variables that holds column values
raw = [links_list] if len(links_list) == 1 else links_list.split('\t')
folders = ['Directories']
filenames = ['Original filenames']
extensions = ['Extensions']

# Final length of col D
final_length = 0

# Function that uses scandir() function to return file link and with depth parameter to control how far the script should scan
def scan_directory(link, depth, reject_bool, reject_pattern):
    if depth == 0:
        return
    with os.scandir(link) as entries:
        for entry in entries:
            if entry.is_file() and (not reject_bool or (reject_bool and not reject_pattern.match(entry.name))):
                folders.append(link)
                filenames.append(os.path.splitext(entry.name)[0])
                extensions.append(os.path.splitext(entry.name)[1])
            elif entry.is_dir():
                scan_directory(entry.path, depth - 1, reject_bool, reject_pattern)

# Incorporate dirs paths, filenames and extensions in arrays
for link in raw:
    if os.path.isdir(link):
        scan_directory(link, depth, reject_bool, reject_pattern)
    elif os.path.isfile(link):
        folders.append(os.path.dirname(link))
        filenames.append(os.path.splitext(os.path.basename(link))[0])
        extensions.append(os.path.splitext(os.path.basename(link))[1])

# Use openpyxl to copy model.xlsx and load values within arrays
wb = openpyxl.load_workbook('model.xlsx', data_only=True)
ws = wb['Terminal']

for i, (fold, name, ext) in enumerate(zip(folders, filenames, extensions), start=1):
    ws.cell(row=i, column=1).value = fold
    ws.cell(row=i, column=2).value = name
    ws.cell(row=i, column=3).value = ext

# Modify width of columns A, B, C, D
for column in ['A', 'B', 'C']:
    length = max(len(str(cell.value)) for cell in ws[column])
    ws.column_dimensions[column].width = length
    if column in ['B', 'C']:
        final_length += length
ws.column_dimensions['D'].width = final_length

# Save and close workbook + check for errors when opening the file
wb.save(temp_file)
wb.close()

result = subprocess.call(['osascript', '-e', 'tell application "Finder" to open POSIX file "{}"'.format(temp_file)])
if result == 0:
     print('File created !,You can now edit new filenames in column D ✍️')
else:
    print('Oups...,There was a problem, please retry ❌️')